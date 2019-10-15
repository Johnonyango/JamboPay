import csv,io
from django.shortcuts import render,redirect
from rest_framework.response import Response
from rest_framework.views import APIView
from django.contrib.auth.decorators import login_required
from .models import *
from .serializer import *
from .permissions import IsAdminOrReadOnly
from rest_framework import status
import requests
# from .forms import *
from django.http import HttpResponse,Http404,HttpResponseRedirect
# from .emails import *
import openpyxl
from rest_framework.permissions import IsAuthenticated  
from django.core.mail import EmailMessage
from django.contrib import messages
from jamboAdmin import views as jamboAdmin_views
from universal_billing_system.emails import *
from universal_billing_system.forms import *
# login
def login(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            user = form.save()
            auth_login(request, user)
            return redirect('index')
    else:
        form = LoginForm()

    return render(request, 'registration/login.html', {'form': form})


@login_required(login_url='register')
def logout_view(request):
    logout(request)
    return redirect('login')


@login_required
def profile(request):
    user = request.user
    images = Image.objects.filter(author=user.profile)
    context = {
        'user': user,
        'images': images
    }

    return render(request, 'timeline/profile.html', context)


# Create your views here.
# def index(request):
#     url = ('jpaye.herokuap.com/api/GetMerchants/')
#     response = requests.get(url)
#     print(response)
   
# def index(request):
#     return render(request, 'index.html')

def index(request):
    current_user=request.user
    if current_user.is_superuser==True:

        print(current_user.is_superuser,"Admiiiiin")

        return redirect(jamboAdmin_views.indexone)
    else:

        return render(request, 'index.html')


@login_required
def upload(request):
    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            filehandle = request.FILES['file']
            return excel.make_response(filehandle.get_sheet(), "csv")
    else:
        form = UploadFileForm()
    return render_to_response('index.html', {'form': form}, context_instance=RequestContext(request))


def bills(request):
    return render(request, 'bills.html')


class MerchantList(APIView):
    permission_classes = (IsAuthenticated,)            # <-- And here

    def get(self, request, format=None):
        permission_classes = (IsAdminOrReadOnly,)
        all_merchants = Merchant.objects.all()
        serializers = MerchantSerializer(all_merchants, many=True)
        return Response(serializers.data)

class RevenueStreamsList(APIView):
    permission_classes = (IsAuthenticated,)            # <-- And here

    def get(self, request, format=None):
        permission_classes = (IsAdminOrReadOnly,)
        all_revenue_streams = Revstreams.objects.all()
        serializers = RevenueStreamsSerializer(all_revenue_streams, many=True)
        return Response(serializers.data)


class GenerateBill(APIView):
    permission_classes = (IsAuthenticated,)            # <-- And here

    # def get(self, request, format=None):
    #     all_bills = Bills.objects.all()
    #     serializers = GenerateBillSerializer(all_bills, many=True)
    #     return Response(serializers.data)
    
    def post(self, request, format=None):
        serializers = BillSerializer(data=request.data)
        if serializers.is_valid():
            serializers.save()
            return Response(serializers.data, status=status.HTTP_201_CREATED)
        return Response(serializers.errors, status=status.HTTP_400_BAD_REQUEST)


class BillsDetails(APIView):
    permission_classes = (IsAuthenticated,)            # <-- And here

    def get(self, request, format=None):
        permission_classes = (IsAdminOrReadOnly,)
        all_bills = Bills.objects.all()
        serializers = BillSerializer(all_bills, many=True)
        return Response(serializers.data)


class GetBillDetails(APIView):
    permission_classes = (IsAuthenticated,)            # <-- And here

    def get_bill(self, pk):
        try:
            return Bills.objects.get(pk=pk)
        except Bills.DoesNotExist:
            return Http404

    def get(self, request, pk, format=None):
        bill = self.get_bill(pk)
        serializers = BillSerializer(bill)
        return Response(serializers.data)


class GetPayments(APIView):
    permission_classes = (IsAuthenticated,)            # <-- And here

    def get(self, request, format=None):
        all_bills = Payments.objects.all()
        serializers = PaymentsSerializer(all_bills, many=True)
        return Response(serializers.data)

    def post(self, request, format=None):
        serializers = PaymentsSerializer(data=request.data)
        if serializers.is_valid():
            serializers.save()
            return Response(serializers.data, status=status.HTTP_201_CREATED)

        # update bills
        specific_bill = Bills.pk
        paid_bill = Payments.bill_number

        if specific_bill == paid_bill:
            specific_bill.status = 1
            # print('true')
            specific_bill.save()

        return Response(serializers.errors, status=status.HTTP_400_BAD_REQUEST)
        permission_classes = (IsAdminOrReadOnly,)


@login_required(login_url='/accounts/login/')
def customers(request):
    url = ('http://127.0.0.1:8000/api/BillsDetails')
    headers = {'Authorization': 'Token b76be7fe9c4ecd62b0e003661426ccbe6cd01d05'}
    response = requests.get(url,headers=headers)
    details = response.json()
    for detail in details:
        Business_name = detail.get('Business_name')
        Email = detail.get('Email')
        Phone_number = detail.get('Phone_number')
        Address = detail.get('Physical_address')
        Code = detail.get('Post_code')
        Town = detail.get('Town')
        Pay_bill = detail.get('JP_paybill')
        Industry = detail.get('Industry')
    return render(request, 'customers.html', {'details': details})


@login_required(login_url='/accounts/login/')
def new_bill(request):
    current_user = request.user
    if request.method == "POST":
        form = BillsForm(request.POST, request.FILES)
        if form.is_valid():
            bill = form.save(commit=False)
            bill.generated_by=current_user
            bill.save()

        # if request.method=="POST":
        # form =BillsForm(request.POST)
        # if form.is_valid():
            name = form.cleaned_data.get('customer_name')
            email = form.cleaned_data.get('customer_email')
            amount = form.cleaned_data.get('amount')
            quantity = form.cleaned_data.get('quantity')

        #     name = request.POST.get('customer_name')
        #     email = request.POST.get('customer_email')
            recipient = NewsLetterRecipients(name=name, email=email,amount=amount,quantity=quantity)
            recipient.save()
            send_notification(name, email,amount=amount,quantity=quantity)
            # recipient = NewsLetterRecipients(name = name,email =email)
            # send_notification(name = name, email = email)


        return HttpResponseRedirect('/index')

    else:
        form = BillsForm()

    return render(request, 'bills/new-bill.html', {"form": form})


def notification(request):
    if request.method == 'POST':
        form = NoteForm(request.POST)
        if form.is_valid():
            name = form.cleaned_data['name']
            email = form.cleaned_data['email']
            recipient = NewsLetterRecipients(name = name,email =email)
            recipient.save()
            send_notification(name = name, email = email)


def upload(request):
    if "GET" == request.method:
        return render(request, 'upload.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        # print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        # print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        # print(active_sheet)

        # reading a cell
        print(worksheet["A1"].value)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)

        return render(request, 'upload.html', {"excel_data": excel_data})




def notification(request):
    if request.method == 'POST':
        form = NoteForm(request.POST)
        if form.is_valid():
            name = form.cleaned_data['name']
            email = form.cleaned_data['email']
            recipient = NewsLetterRecipients(name=name, email=email)
            recipient.save()
            send_notification(name=name, email=email)
    else:
        form = NoteForm()
    return render(request, 'note.html', {'form': form})

@login_required(login_url='/accounts/login/')
def uploadCSV(request):
    current_user = request.user
    template = "bills_upload.html"
    prompt = {"order":"order of csv should be as follows: \n customer_name,customer_phone,customer_email,narration,amount,quantity,post_date,due_date"}
    if request.method == "GET":
        return render(request, template, prompt)
    csv_file = request.FILES['file']
    if not csv_file.name.endswith('.csv'):
        message.error(request,"this is not a csv file")
    else:
        messages.success(request,'Upload successfull')

    data_set = csv_file.read().decode('UTF-8')
    io_string = io.StringIO(data_set)
    next(io_string)
    for column in csv.reader(io_string, delimiter=',', quotechar="|"):
        _, created = Bills.objects.update_or_create(
            customer_name=column[0],
            customer_phone=column[1],
            customer_email=column[2],
            # Revstreams =column[3],
            narration=column[3],
            amount=column[4],
            quantity=column[5],
            post_date=column[6],
            # status=column[8],
            due_date=column[7],
            generated_by=current_user
        )
    context = {}
    return render(request, template, context)


@login_required(login_url='/accounts/login/')
def search_results(request):
    current_user = request.user
    if 'customer_name' in request.GET and request.GET["customer_name"]:
        search_term = request.GET.get("customer_name")
        names = Bills.search_by_name(search_term)
        message = f"{search_term}"

        print(names)

        return render(request, 'search.html', {"message": message, "names": names})

    else:
        message = "You haven't searched for any term."
        return render(request, 'search.html', {"message": message})



@login_required(login_url='/accounts/login/')
def merchant_bills(request):
    details = Bills.get_merchant_bills(request.user)

    message = f"The Following  bills were generated by : {request.user}"


    return render(request, 'mybills.html', {'details': details,'message':message})



@login_required
def addEmployee(request):
    if request.method == 'POST':
        form = AddEmployeeForm(request.POST)
        if form.is_valid():
            
            name=form.cleaned_data.get('username')
            email = form.cleaned_data.get('email')
            password = form.cleaned_data.get('password1') 
            form.save()

            welcome_email(name,email, password)

            messages.success(request,'Employee added succesfully')
            return redirect('Index')

    else:

        form=AddEmployeeForm()

    return render(request, 'admins/add_employee.html', {'form': form})



#template
import mimetypes

